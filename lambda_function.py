import json

# def lambda_handler(event, context):
#     print("Received event: " + json.dumps(event, indent=2))
#     # TODO implement
#     message = 'Hello {} {}!'.format(event['first_name'], 
#                                     event['last_name'])  
#     return {
#         'statusCode': 200,
#         'body': message
#     }
from botocore.client import Config
import boto3
from botocore.exceptions import ClientError
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from fpdf import FPDF
import logging

def lambda_handler(event, context):
    print("Received event: " + json.dumps(event, indent=2))
    # TODO implement
    message = 'Hello {} {}!'.format(event['first_name'], 
                                    event['last_name'])  
    #runExcel("test.xlsx")
    runPdf("test2.pdf")
    #print("output")
    return {
        'statusCode': 200,
        'body': message
    }

def runPdf(fileName):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('arial', 'B', 13.0)
    pdf.cell(ln=0, h=5.0, align='L', w=0, txt="Hello", border=0)
    pdfOut = pdf.output(name=fileName, dest='S').encode('latin-1')
    output = BytesIO(pdfOut)
    print(output)
    #config = Config(connect_timeout=5, read_timeout=5)
    
    s3 = boto3.resource(
        's3',
        region_name='us-west-2'
        #aws_access_key_id='',
        #aws_secret_access_key=''
    )
    #s3.Bucket('awsserverless-robk').upload_fileobj(output, fileName)
    print("before try")
    try:
        target_object = s3.Object('awsserverless-robk', fileName)
        target_object.put(Body=output)
    except ClientError as e:
        logging.error(e)
        print(e)   

def runExcel(fileName):
    wb = Workbook()
    #with open(fileName, 'wb') as file:
    ws1 = wb.active
    #print("file Name: " + file.name)
    wb = Workbook()
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
        
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    #output = BytesIO(file)
    #wb.save(file)
    output = BytesIO(save_virtual_workbook(wb))
    # with NamedTemporaryFile(delete=False) as tmp:
    #     wb.save(tmp.name)
    #     output = BytesIO(tmp.read())

    s3 = boto3.resource(
        's3',
        region_name='us-west-2',
        #aws_access_key_id='',
        #aws_secret_access_key=''
    )
    #s3.Bucket('awsserverless-robk').upload_fileobj(output, fileName)
    try:
        target_object = s3.Object('awsserverless-robk',fileName)
        target_object.put(Body=output)
    except ClientError as e:
        logging.error(e)  
        #fileobj = BytesIO(file.read())
        #s3.upload_fileobj(output, 'awsserverless-robk', fileName)